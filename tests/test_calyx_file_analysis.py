from __future__ import annotations

import io

import pytest
from openpyxl import Workbook

from app.calyx_conversation.file_analysis import chart_spec, parse_tabular_file
from app.calyx_conversation.routes import DatasetAnalysisRequest, run_dataset_analysis


def test_parse_csv_normalizes_headers_rows_and_numeric_types():
    parsed = parse_tabular_file(
        b"species,temp,temp\nA,10,11\nB,12.5,13\n",
        "orchids.csv",
    )
    assert parsed["row_count"] == 2
    assert list(parsed["columns"]) == ["species", "temp", "temp_2"]
    assert parsed["columns"]["temp"] == [10, 12.5]
    assert parsed["columns"]["temp_2"] == [11, 13]
    assert parsed["metadata"]["format"] == "csv"


def test_parse_xlsx_selects_requested_sheet():
    workbook = Workbook()
    active = workbook.active
    active.title = "Summary"
    active.append(["species", "value"])
    active.append(["A", 1])
    detail = workbook.create_sheet("Detail")
    detail.append(["species", "value"])
    detail.append(["B", 9])
    payload = io.BytesIO()
    workbook.save(payload)
    workbook.close()

    parsed = parse_tabular_file(payload.getvalue(), "orchids.xlsx", sheet_name="Detail")
    assert parsed["row_count"] == 1
    assert parsed["columns"]["species"] == ["B"]
    assert parsed["columns"]["value"] == [9]
    assert parsed["metadata"]["sheet"] == "Detail"


def test_parse_xlsx_rejects_unknown_sheet():
    workbook = Workbook()
    workbook.active.append(["x"])
    payload = io.BytesIO()
    workbook.save(payload)
    workbook.close()
    with pytest.raises(ValueError, match="worksheet not found"):
        parse_tabular_file(payload.getvalue(), "data.xlsx", sheet_name="Missing")


def test_parse_rejects_unsupported_extension():
    with pytest.raises(ValueError, match="CSV and XLSX"):
        parse_tabular_file(b"x,y\n1,2\n", "data.tsv")


def test_uploaded_csv_columns_feed_dataset_analysis_directly():
    parsed = parse_tabular_file(b"x,y\n1,2\n2,4\n3,6\n", "data.csv")
    result = run_dataset_analysis(
        DatasetAnalysisRequest(operation="correlation_matrix", columns=parsed["columns"])
    )
    assert result["matrix"]["x"]["y"] == pytest.approx(1.0)


def test_chart_spec_validates_columns_shape_and_inline_data():
    columns = {"month": [1, 2, 3], "flowers": [3, 5, 8]}
    spec = chart_spec(columns, chart_type="line", x="month", y="flowers")
    assert spec["chart_type"] == "line"
    assert spec["row_count"] == 3
    assert spec["x"] == "month"
    assert spec["y"] == "flowers"
    assert spec["data"] == [
        {"x": 1, "y": 3},
        {"x": 2, "y": 5},
        {"x": 3, "y": 8},
    ]
    assert spec["data_inline"] is True


def test_histogram_chart_spec_filters_non_numeric_values():
    spec = chart_spec(
        {"temperature": [10, None, "bad", 12.5]},
        chart_type="histogram",
        x="temperature",
    )
    assert spec["data"] == [{"value": 10}, {"value": 12.5}]


def test_chart_spec_rejects_missing_axis():
    with pytest.raises(ValueError, match="requires y"):
        chart_spec({"x": [1]}, chart_type="scatter", x="x")
