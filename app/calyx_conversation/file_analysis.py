from __future__ import annotations

import csv
import io
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

MAX_UPLOAD_BYTES = 10 * 1024 * 1024
MAX_ROWS = 100_000
MAX_COLUMNS = 500
SUPPORTED_EXTENSIONS = {".csv", ".xlsx"}


def _normalize_headers(values: list[Any]) -> list[str]:
    headers: list[str] = []
    seen: dict[str, int] = {}
    for index, value in enumerate(values, start=1):
        base = str(value).strip() if value is not None else ""
        base = base or f"column_{index}"
        count = seen.get(base, 0) + 1
        seen[base] = count
        headers.append(base if count == 1 else f"{base}_{count}")
    if len(headers) > MAX_COLUMNS:
        raise ValueError(f"dataset exceeds {MAX_COLUMNS} columns")
    return headers


def _columns_from_rows(rows: list[list[Any]]) -> dict[str, list[Any]]:
    if not rows:
        raise ValueError("dataset is empty")
    headers = _normalize_headers(rows[0])
    data_rows = rows[1:]
    if len(data_rows) > MAX_ROWS:
        raise ValueError(f"dataset exceeds {MAX_ROWS} data rows")
    columns = {header: [] for header in headers}
    for row in data_rows:
        normalized = list(row[: len(headers)]) + [None] * max(0, len(headers) - len(row))
        for header, value in zip(headers, normalized, strict=True):
            columns[header].append(value)
    return columns


def _parse_csv(content: bytes) -> tuple[dict[str, list[Any]], dict[str, Any]]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise ValueError("CSV must be UTF-8 encoded") from exc
    reader = csv.reader(io.StringIO(text))
    rows: list[list[Any]] = []
    for index, row in enumerate(reader):
        if index > MAX_ROWS:
            raise ValueError(f"dataset exceeds {MAX_ROWS} data rows")
        if len(row) > MAX_COLUMNS:
            raise ValueError(f"dataset exceeds {MAX_COLUMNS} columns")
        rows.append(row)
    columns = _columns_from_rows(rows)
    return columns, {"format": "csv", "sheet": None}


def _parse_xlsx(content: bytes, sheet_name: str | None) -> tuple[dict[str, list[Any]], dict[str, Any]]:
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError("unable to read XLSX workbook") from exc
    try:
        if not workbook.sheetnames:
            raise ValueError("workbook contains no worksheets")
        selected = sheet_name or workbook.sheetnames[0]
        if selected not in workbook.sheetnames:
            raise ValueError(f"worksheet not found: {selected}")
        worksheet = workbook[selected]
        rows: list[list[Any]] = []
        for index, row in enumerate(worksheet.iter_rows(values_only=True)):
            if index > MAX_ROWS:
                raise ValueError(f"dataset exceeds {MAX_ROWS} data rows")
            values = list(row)
            if len(values) > MAX_COLUMNS:
                raise ValueError(f"dataset exceeds {MAX_COLUMNS} columns")
            rows.append(values)
        columns = _columns_from_rows(rows)
        return columns, {"format": "xlsx", "sheet": selected, "available_sheets": workbook.sheetnames}
    finally:
        workbook.close()


def parse_tabular_file(content: bytes, filename: str, *, sheet_name: str | None = None) -> dict[str, Any]:
    if not content:
        raise ValueError("uploaded file is empty")
    if len(content) > MAX_UPLOAD_BYTES:
        raise ValueError(f"uploaded file exceeds {MAX_UPLOAD_BYTES // (1024 * 1024)} MB")
    extension = Path(filename or "").suffix.casefold()
    if extension not in SUPPORTED_EXTENSIONS:
        raise ValueError("supported upload formats are CSV and XLSX")

    if extension == ".csv":
        columns, metadata = _parse_csv(content)
    else:
        columns, metadata = _parse_xlsx(content, sheet_name)

    row_count = len(next(iter(columns.values()))) if columns else 0
    return {
        "filename": Path(filename).name,
        "row_count": row_count,
        "column_count": len(columns),
        "columns": columns,
        "metadata": metadata,
    }


def chart_spec(columns: dict[str, list[Any]], *, chart_type: str, x: str | None = None, y: str | None = None) -> dict[str, Any]:
    supported = {"scatter", "line", "bar", "histogram"}
    if chart_type not in supported:
        raise ValueError(f"chart_type must be one of {', '.join(sorted(supported))}")
    if x and x not in columns:
        raise ValueError(f"unknown x column: {x}")
    if y and y not in columns:
        raise ValueError(f"unknown y column: {y}")
    if chart_type in {"scatter", "line", "bar"} and not x:
        raise ValueError(f"{chart_type} chart requires x column")
    if chart_type in {"scatter", "line", "bar"} and not y:
        raise ValueError(f"{chart_type} chart requires y column")
    if chart_type == "histogram" and not (x or y):
        raise ValueError("histogram requires a numeric column in x or y")
    return {
        "version": "calyx-chart-spec-1",
        "chart_type": chart_type,
        "x": x,
        "y": y,
        "row_count": len(next(iter(columns.values()))) if columns else 0,
        "rendering": "frontend",
        "data_inline": False,
    }
