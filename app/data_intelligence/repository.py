from __future__ import annotations

import csv
import hashlib
import io
import json
import re
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from threading import RLock
from typing import Any

from openpyxl import load_workbook

from .models import DataIntelligenceError

_SAFE = re.compile(r"^[A-Za-z0-9._-]+$")


@dataclass(frozen=True, slots=True)
class DatasetVersion:
    dataset_id: str
    version_id: str
    owner: str
    project_id: str
    logical_name: str
    format: str
    content_hash: str
    byte_size: int
    created_at: str

    def to_dict(self) -> dict[str, Any]:
        return {
            "dataset_id": self.dataset_id,
            "version_id": self.version_id,
            "owner": self.owner,
            "project_id": self.project_id,
            "logical_name": self.logical_name,
            "format": self.format,
            "content_hash": self.content_hash,
            "byte_size": self.byte_size,
            "created_at": self.created_at,
        }


class FileDatasetRepository:
    """Tenant/project-scoped, content-addressed dataset and analysis storage."""

    def __init__(self, root: str | Path) -> None:
        self.root = Path(root)
        self._lock = RLock()

    def _part(self, value: str, code: str) -> str:
        value = value.strip()
        if not value or not _SAFE.match(value) or value in {".", ".."}:
            raise DataIntelligenceError(code)
        return value

    def _scope(self, owner: str, project_id: str) -> Path:
        return self.root / self._part(owner, "INVALID_OWNER") / self._part(project_id, "INVALID_PROJECT_ID")

    def _version_dir(self, owner: str, project_id: str, dataset_id: str, version_id: str) -> Path:
        return (
            self._scope(owner, project_id)
            / self._part(dataset_id, "INVALID_DATASET_ID")
            / self._part(version_id, "INVALID_VERSION_ID")
        )

    @staticmethod
    def _hash(data: bytes) -> str:
        return hashlib.sha256(data).hexdigest()

    @staticmethod
    def _dataset_id(owner: str, project_id: str, logical_name: str) -> str:
        payload = f"{owner}\x1f{project_id}\x1f{logical_name.strip().casefold()}".encode("utf-8")
        return hashlib.sha256(payload).hexdigest()[:32]

    def ingest(
        self,
        *,
        owner: str,
        project_id: str,
        logical_name: str,
        filename: str,
        data: bytes,
    ) -> tuple[DatasetVersion, bool]:
        suffix = Path(filename).suffix.lower()
        if suffix not in {".csv", ".xlsx"}:
            raise DataIntelligenceError("UNSUPPORTED_DATASET_FORMAT", {"suffix": suffix})
        if not data:
            raise DataIntelligenceError("EMPTY_DATASET")
        dataset_id = self._dataset_id(owner, project_id, logical_name)
        version_id = self._hash(data)
        directory = self._version_dir(owner, project_id, dataset_id, version_id)
        metadata_path = directory / "dataset.json"
        with self._lock:
            if metadata_path.is_file():
                return self.get(owner, project_id, dataset_id, version_id), False
            directory.mkdir(parents=True, exist_ok=True)
            raw_path = directory / f"source{suffix}"
            temp = raw_path.with_suffix(raw_path.suffix + ".tmp")
            temp.write_bytes(data)
            temp.replace(raw_path)
            metadata = DatasetVersion(
                dataset_id=dataset_id,
                version_id=version_id,
                owner=owner,
                project_id=project_id,
                logical_name=logical_name.strip(),
                format=suffix[1:],
                content_hash=version_id,
                byte_size=len(data),
                created_at=datetime.now(timezone.utc).isoformat(),
            )
            self._write_json(metadata_path, metadata.to_dict())
            return metadata, True

    def get(self, owner: str, project_id: str, dataset_id: str, version_id: str) -> DatasetVersion:
        path = self._version_dir(owner, project_id, dataset_id, version_id) / "dataset.json"
        if not path.is_file():
            raise DataIntelligenceError("DATASET_VERSION_NOT_FOUND")
        return DatasetVersion(**json.loads(path.read_text(encoding="utf-8")))

    def read_rows(
        self, owner: str, project_id: str, dataset_id: str, version_id: str
    ) -> tuple[list[str], list[dict[str, Any]]]:
        metadata = self.get(owner, project_id, dataset_id, version_id)
        directory = self._version_dir(owner, project_id, dataset_id, version_id)
        if metadata.format == "csv":
            raw = (directory / "source.csv").read_bytes()
            try:
                text = raw.decode("utf-8-sig")
            except UnicodeDecodeError as exc:
                raise DataIntelligenceError("CSV_MUST_BE_UTF8") from exc
            reader = csv.DictReader(io.StringIO(text, newline=""))
            if not reader.fieldnames:
                raise DataIntelligenceError("DATASET_HEADER_REQUIRED")
            columns = [str(value).strip() for value in reader.fieldnames]
            if any(not value for value in columns) or len(set(columns)) != len(columns):
                raise DataIntelligenceError("DATASET_COLUMNS_INVALID")
            return columns, [dict(row) for row in reader]
        workbook = load_workbook(directory / "source.xlsx", read_only=True, data_only=True)
        try:
            sheet = workbook.active
            iterator = sheet.iter_rows(values_only=True)
            header = next(iterator, None)
            if not header:
                raise DataIntelligenceError("DATASET_HEADER_REQUIRED")
            columns = [str(value).strip() if value is not None else "" for value in header]
            if any(not value for value in columns) or len(set(columns)) != len(columns):
                raise DataIntelligenceError("DATASET_COLUMNS_INVALID")
            rows = [
                {columns[index]: row[index] if index < len(row) else None for index in range(len(columns))}
                for row in iterator
            ]
            return columns, rows
        finally:
            workbook.close()

    def profile_path(self, owner: str, project_id: str, dataset_id: str, version_id: str) -> Path:
        return self._version_dir(owner, project_id, dataset_id, version_id) / "profile.json"

    def save_profile(self, owner: str, project_id: str, dataset_id: str, version_id: str, profile: dict[str, Any]) -> None:
        self._write_json(self.profile_path(owner, project_id, dataset_id, version_id), profile)

    def get_profile(self, owner: str, project_id: str, dataset_id: str, version_id: str) -> dict[str, Any] | None:
        path = self.profile_path(owner, project_id, dataset_id, version_id)
        return json.loads(path.read_text(encoding="utf-8")) if path.is_file() else None

    def analysis_dir(self, owner: str, project_id: str, dataset_id: str, version_id: str, analysis_id: str) -> Path:
        return self._version_dir(owner, project_id, dataset_id, version_id) / "analyses" / self._part(analysis_id, "INVALID_ANALYSIS_ID")

    def save_analysis(
        self,
        *,
        owner: str,
        project_id: str,
        dataset_id: str,
        version_id: str,
        analysis_id: str,
        manifest: dict[str, Any],
        table_bytes: bytes,
        chart_bytes: bytes | None,
    ) -> dict[str, str]:
        directory = self.analysis_dir(owner, project_id, dataset_id, version_id, analysis_id)
        directory.mkdir(parents=True, exist_ok=True)
        artifacts = {"table.json": self._write_bytes(directory / "table.json", table_bytes)}
        if chart_bytes is not None:
            artifacts["chart.svg"] = self._write_bytes(directory / "chart.svg", chart_bytes)
        final_manifest = {**manifest, "artifact_hashes": artifacts}
        self._write_json(directory / "manifest.json", final_manifest)
        return artifacts

    def get_analysis(self, owner: str, project_id: str, dataset_id: str, version_id: str, analysis_id: str) -> dict[str, Any]:
        path = self.analysis_dir(owner, project_id, dataset_id, version_id, analysis_id) / "manifest.json"
        if not path.is_file():
            raise DataIntelligenceError("ANALYSIS_NOT_FOUND")
        return json.loads(path.read_text(encoding="utf-8"))

    @staticmethod
    def _write_bytes(path: Path, data: bytes) -> str:
        temp = path.with_suffix(path.suffix + ".tmp")
        temp.write_bytes(data)
        temp.replace(path)
        return hashlib.sha256(data).hexdigest()

    @classmethod
    def _write_json(cls, path: Path, payload: dict[str, Any]) -> None:
        data = (json.dumps(payload, indent=2, sort_keys=True, ensure_ascii=False, default=str) + "\n").encode("utf-8")
        cls._write_bytes(path, data)
