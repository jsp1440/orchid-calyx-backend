import io
from datetime import datetime
from typing import List, Optional

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from fastapi import APIRouter, Depends, HTTPException, Query, Response, UploadFile, File
from sqlalchemy import select, func
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.deps import get_db
from app.models import (
    Show,
    Volunteer,
    VolunteerRole,
    VolunteerShift,
    VolunteerAssignment,
)
from app.schemas import (
    VolunteerRoleCreate,
    VolunteerRoleOut,
    VolunteerShiftCreate,
    VolunteerShiftOut,
    VolunteerCreate,
    VolunteerOut,
    VolunteerAssignmentCreate,
    VolunteerAssignmentOut,
    VolunteerAssignmentMove,
)
from app.security import verify_api_key

router = APIRouter(
    prefix="/api/shows/{show_id}/volunteer",
    tags=["Volunteer Operations"],
    dependencies=[Depends(verify_api_key)],
)

VALID_ASSIGNMENT_STATUSES = {"assigned", "confirmed", "checked_in", "no_show"}


def _require_show(db: Session, show_id: str) -> Show:
    show = db.get(Show, show_id)
    if not show:
        raise HTTPException(status_code=404, detail="Show not found")
    return show


# ── Roles ──────────────────────────────────────────────────────────

@router.post("/roles", response_model=VolunteerRoleOut)
def create_role(show_id: str, data: VolunteerRoleCreate, db: Session = Depends(get_db)):
    _require_show(db, show_id)
    role = VolunteerRole(show_id=show_id, **data.model_dump())
    db.add(role)
    db.commit()
    db.refresh(role)
    return role


@router.get("/roles", response_model=List[VolunteerRoleOut])
def list_roles(show_id: str, db: Session = Depends(get_db)):
    return db.execute(select(VolunteerRole).where(VolunteerRole.show_id == show_id)).scalars().all()


# ── Shifts ─────────────────────────────────────────────────────────

@router.post("/shifts", response_model=VolunteerShiftOut)
def create_shift(show_id: str, data: VolunteerShiftCreate, db: Session = Depends(get_db)):
    _require_show(db, show_id)
    role = db.get(VolunteerRole, data.role_id)
    if not role or role.show_id != show_id:
        raise HTTPException(status_code=404, detail="Role not found for this show")

    shift = VolunteerShift(
        show_id=show_id,
        role_id=data.role_id,
        start_time=data.start_time,
        end_time=data.end_time,
        capacity=data.capacity or 1,
    )
    db.add(shift)
    db.commit()
    db.refresh(shift)
    return shift


@router.get("/shifts", response_model=List[VolunteerShiftOut])
def list_shifts(show_id: str, db: Session = Depends(get_db)):
    return db.execute(
        select(VolunteerShift).where(VolunteerShift.show_id == show_id)
    ).scalars().all()


@router.patch("/shifts/{shift_id}", response_model=VolunteerShiftOut)
def update_shift(show_id: str, shift_id: str, db: Session = Depends(get_db),
                 role_id: Optional[str] = Query(None),
                 start_time: Optional[datetime] = Query(None),
                 end_time: Optional[datetime] = Query(None),
                 capacity: Optional[int] = Query(None)):
    shift = db.get(VolunteerShift, shift_id)
    if not shift or shift.show_id != show_id:
        raise HTTPException(status_code=404, detail="Shift not found")
    if role_id is not None:
        shift.role_id = role_id
    if start_time is not None:
        shift.start_time = start_time
    if end_time is not None:
        shift.end_time = end_time
    if capacity is not None:
        shift.capacity = capacity
    db.commit()
    db.refresh(shift)
    return shift


@router.delete("/shifts/{shift_id}")
def delete_shift(show_id: str, shift_id: str, db: Session = Depends(get_db)):
    shift = db.get(VolunteerShift, shift_id)
    if not shift or shift.show_id != show_id:
        raise HTTPException(status_code=404, detail="Shift not found")
    db.delete(shift)
    db.commit()
    return {"status": "deleted"}


# ── Volunteers ─────────────────────────────────────────────────────

@router.post("/volunteers", response_model=VolunteerOut)
def create_volunteer(show_id: str, data: VolunteerCreate, db: Session = Depends(get_db)):
    _require_show(db, show_id)
    vol = Volunteer(show_id=show_id, **data.model_dump())
    db.add(vol)
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=409, detail="Volunteer with this email already exists for this show")
    db.refresh(vol)
    return vol


@router.get("/volunteers", response_model=List[VolunteerOut])
def list_volunteers(show_id: str, db: Session = Depends(get_db)):
    return db.execute(select(Volunteer).where(Volunteer.show_id == show_id)).scalars().all()


@router.patch("/volunteers/{volunteer_id}", response_model=VolunteerOut)
def update_volunteer(show_id: str, volunteer_id: str, db: Session = Depends(get_db),
                     name: Optional[str] = Query(None),
                     phone: Optional[str] = Query(None),
                     opt_in_sms: Optional[bool] = Query(None),
                     notes: Optional[str] = Query(None),
                     approved: Optional[bool] = Query(None)):
    vol = db.get(Volunteer, volunteer_id)
    if not vol or vol.show_id != show_id:
        raise HTTPException(status_code=404, detail="Volunteer not found")
    if name is not None:
        vol.name = name
    if phone is not None:
        vol.phone = phone
    if opt_in_sms is not None:
        vol.opt_in_sms = opt_in_sms
    if notes is not None:
        vol.notes = notes
    if approved is not None:
        vol.approved = approved
    db.commit()
    db.refresh(vol)
    return vol


# ── Assignments ────────────────────────────────────────────────────

@router.post("/assignments", response_model=VolunteerAssignmentOut)
def create_assignment(show_id: str, data: VolunteerAssignmentCreate, db: Session = Depends(get_db)):
    _require_show(db, show_id)

    vol = db.get(Volunteer, data.volunteer_id)
    if not vol or vol.show_id != show_id:
        raise HTTPException(status_code=404, detail="Volunteer not found for this show")

    shift = db.get(VolunteerShift, data.shift_id)
    if not shift or shift.show_id != show_id:
        raise HTTPException(status_code=404, detail="Shift not found for this show")

    if data.status and data.status not in VALID_ASSIGNMENT_STATUSES:
        raise HTTPException(status_code=422, detail=f"Invalid status. Must be one of: {', '.join(sorted(VALID_ASSIGNMENT_STATUSES))}")

    current = db.execute(
        select(func.count(VolunteerAssignment.id)).where(
            VolunteerAssignment.shift_id == data.shift_id,
        )
    ).scalar_one()
    if current >= shift.capacity:
        raise HTTPException(status_code=409, detail="Shift is full")

    assignment = VolunteerAssignment(
        show_id=show_id,
        volunteer_id=data.volunteer_id,
        shift_id=data.shift_id,
        status=data.status or "assigned",
    )
    db.add(assignment)
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=409, detail="Volunteer already assigned to this shift")
    db.refresh(assignment)
    return assignment


@router.get("/assignments", response_model=List[VolunteerAssignmentOut])
def list_assignments(show_id: str, db: Session = Depends(get_db)):
    return db.execute(
        select(VolunteerAssignment).where(VolunteerAssignment.show_id == show_id)
    ).scalars().all()


@router.patch("/assignments/{assignment_id}/status", response_model=VolunteerAssignmentOut)
def update_assignment_status(show_id: str, assignment_id: str,
                             status: str = Query(...),
                             db: Session = Depends(get_db)):
    if status not in VALID_ASSIGNMENT_STATUSES:
        raise HTTPException(status_code=422, detail=f"Invalid status. Must be one of: {', '.join(sorted(VALID_ASSIGNMENT_STATUSES))}")
    assignment = db.get(VolunteerAssignment, assignment_id)
    if not assignment or assignment.show_id != show_id:
        raise HTTPException(status_code=404, detail="Assignment not found")
    assignment.status = status
    db.commit()
    db.refresh(assignment)
    return assignment


@router.patch("/assignments/{assignment_id}/move", response_model=VolunteerAssignmentOut)
def move_assignment(show_id: str, assignment_id: str, data: VolunteerAssignmentMove, db: Session = Depends(get_db)):
    assignment = db.get(VolunteerAssignment, assignment_id)
    if not assignment or assignment.show_id != show_id:
        raise HTTPException(status_code=404, detail="Assignment not found")

    new_shift = db.get(VolunteerShift, data.shift_id)
    if not new_shift or new_shift.show_id != show_id:
        raise HTTPException(status_code=404, detail="Target shift not found")

    existing = db.execute(
        select(VolunteerAssignment).where(
            VolunteerAssignment.shift_id == data.shift_id,
            VolunteerAssignment.volunteer_id == assignment.volunteer_id,
        )
    ).scalar_one_or_none()
    if existing:
        raise HTTPException(status_code=409, detail="Volunteer already assigned to target shift")

    current = db.execute(
        select(func.count(VolunteerAssignment.id)).where(
            VolunteerAssignment.shift_id == data.shift_id,
        )
    ).scalar_one()
    if current >= new_shift.capacity:
        raise HTTPException(status_code=409, detail="Target shift is full")

    assignment.shift_id = data.shift_id
    db.commit()
    db.refresh(assignment)
    return assignment


@router.delete("/assignments/{assignment_id}")
def delete_assignment(show_id: str, assignment_id: str, db: Session = Depends(get_db)):
    assignment = db.get(VolunteerAssignment, assignment_id)
    if not assignment or assignment.show_id != show_id:
        raise HTTPException(status_code=404, detail="Assignment not found")
    db.delete(assignment)
    db.commit()
    return {"status": "deleted"}


# ── Check-in / Check-out (via assignment status) ──────────────────

@router.post("/check-in", response_model=VolunteerAssignmentOut)
def check_in(show_id: str, volunteer_id: str = Query(...), shift_id: str = Query(...),
             method: Optional[str] = Query("web"), db: Session = Depends(get_db)):
    assignment = db.execute(
        select(VolunteerAssignment).where(
            VolunteerAssignment.show_id == show_id,
            VolunteerAssignment.volunteer_id == volunteer_id,
            VolunteerAssignment.shift_id == shift_id,
        )
    ).scalar_one_or_none()

    if not assignment:
        raise HTTPException(status_code=404, detail="Assignment not found")
    if assignment.status == "checked_in":
        raise HTTPException(status_code=409, detail="Already checked in")

    assignment.status = "checked_in"
    db.commit()
    db.refresh(assignment)
    return assignment


# ── Excel Export ───────────────────────────────────────────────────

@router.get("/export.xlsx")
def export_xlsx(show_id: str, db: Session = Depends(get_db)):
    show = _require_show(db, show_id)

    roles = {r.id: r for r in db.execute(select(VolunteerRole).where(VolunteerRole.show_id == show_id)).scalars().all()}
    shifts = {s.id: s for s in db.execute(select(VolunteerShift).where(VolunteerShift.show_id == show_id)).scalars().all()}
    vols = {v.id: v for v in db.execute(select(Volunteer).where(Volunteer.show_id == show_id)).scalars().all()}
    assignments = db.execute(select(VolunteerAssignment).where(VolunteerAssignment.show_id == show_id)).scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Volunteer Schedule"

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = ["volunteer_name", "email", "phone", "approved", "opt_in_sms",
               "notes", "role_name", "shift_start", "shift_end", "capacity",
               "assignment_status"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    row = 2
    if assignments:
        for a in assignments:
            vol = vols.get(a.volunteer_id)
            shift = shifts.get(a.shift_id)
            role = roles.get(shift.role_id) if shift else None
            ws.cell(row=row, column=1, value=vol.name if vol else "")
            ws.cell(row=row, column=2, value=vol.email if vol else "")
            ws.cell(row=row, column=3, value=vol.phone if vol else "")
            ws.cell(row=row, column=4, value=vol.approved if vol else False)
            ws.cell(row=row, column=5, value=vol.opt_in_sms if vol else False)
            ws.cell(row=row, column=6, value=vol.notes if vol else "")
            ws.cell(row=row, column=7, value=role.name if role else "")
            ws.cell(row=row, column=8, value=shift.start_time.isoformat() if shift else "")
            ws.cell(row=row, column=9, value=shift.end_time.isoformat() if shift else "")
            ws.cell(row=row, column=10, value=shift.capacity if shift else 0)
            ws.cell(row=row, column=11, value=a.status)
            row += 1
    else:
        for vol in vols.values():
            ws.cell(row=row, column=1, value=vol.name)
            ws.cell(row=row, column=2, value=vol.email)
            ws.cell(row=row, column=3, value=vol.phone or "")
            ws.cell(row=row, column=4, value=vol.approved)
            ws.cell(row=row, column=5, value=vol.opt_in_sms)
            ws.cell(row=row, column=6, value=vol.notes or "")
            row += 1

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=volunteers_{show_id}.xlsx"},
    )


# ── Excel Import (with conflict detection) ─────────────────────────

@router.post("/import")
async def import_volunteers(show_id: str,
                            override_conflicts: bool = Query(False),
                            file: UploadFile = File(...),
                            db: Session = Depends(get_db)):
    _require_show(db, show_id)

    raw = await file.read()

    if file.filename and file.filename.endswith(".xlsx"):
        rows = _parse_xlsx(raw)
    else:
        rows = _parse_csv(raw)

    role_cache = {}
    for r in db.execute(select(VolunteerRole).where(VolunteerRole.show_id == show_id)).scalars().all():
        role_cache[r.name.strip().lower()] = r

    shift_cache = {}
    for s in db.execute(select(VolunteerShift).where(VolunteerShift.show_id == show_id)).scalars().all():
        key = (s.role_id, s.start_time.isoformat(), s.end_time.isoformat())
        shift_cache[key] = s

    stats = {
        "volunteers_created": 0, "volunteers_updated": 0,
        "roles_created": 0, "shifts_created": 0,
        "assignments_created": 0, "rows_processed": 0,
        "conflicts": [],
    }

    for row in rows:
        stats["rows_processed"] += 1

        vol_name = (row.get("volunteer_name") or "").strip()
        vol_email = (row.get("email") or "").strip()
        vol_phone = (row.get("phone") or "").strip()
        vol_approved = _parse_bool(row.get("approved", ""))
        vol_opt_sms = _parse_bool(row.get("opt_in_sms", ""))
        vol_notes = (row.get("notes") or "").strip()
        role_name = (row.get("role_name") or "").strip()
        s_start = (row.get("shift_start") or "").strip()
        s_end = (row.get("shift_end") or "").strip()
        a_status = (row.get("assignment_status") or "assigned").strip()

        if not vol_name or not vol_email:
            continue

        vol = db.execute(
            select(Volunteer).where(Volunteer.show_id == show_id, Volunteer.email == vol_email)
        ).scalar_one_or_none()

        if vol:
            changes = []
            if vol.name != vol_name:
                changes.append(f"name: {vol.name} -> {vol_name}")
            if vol_phone and vol.phone != vol_phone:
                changes.append(f"phone: {vol.phone} -> {vol_phone}")

            if changes and not override_conflicts:
                stats["conflicts"].append({
                    "email": vol_email,
                    "changes": changes,
                    "action": "skipped (use override_conflicts=true to apply)",
                })
                continue

            vol.name = vol_name
            if vol_phone:
                vol.phone = vol_phone
            vol.approved = vol_approved
            vol.opt_in_sms = vol_opt_sms
            if vol_notes:
                vol.notes = vol_notes
            stats["volunteers_updated"] += 1
        else:
            vol = Volunteer(
                show_id=show_id,
                name=vol_name,
                email=vol_email,
                phone=vol_phone or None,
                approved=vol_approved,
                opt_in_sms=vol_opt_sms,
                notes=vol_notes or None,
            )
            db.add(vol)
            db.flush()
            stats["volunteers_created"] += 1

        if not role_name or not s_start or not s_end:
            continue

        rkey = role_name.lower()
        if rkey not in role_cache:
            role = VolunteerRole(show_id=show_id, name=role_name)
            db.add(role)
            db.flush()
            role_cache[rkey] = role
            stats["roles_created"] += 1
        role = role_cache[rkey]

        try:
            st = datetime.fromisoformat(s_start)
            et = datetime.fromisoformat(s_end)
        except ValueError:
            continue

        skey = (role.id, st.isoformat(), et.isoformat())
        if skey not in shift_cache:
            shift = VolunteerShift(
                show_id=show_id, role_id=role.id,
                start_time=st, end_time=et,
            )
            db.add(shift)
            db.flush()
            shift_cache[skey] = shift
            stats["shifts_created"] += 1
        shift = shift_cache[skey]

        if a_status not in VALID_ASSIGNMENT_STATUSES:
            a_status = "assigned"

        existing_a = db.execute(
            select(VolunteerAssignment).where(
                VolunteerAssignment.shift_id == shift.id,
                VolunteerAssignment.volunteer_id == vol.id,
            )
        ).scalar_one_or_none()
        if not existing_a:
            a = VolunteerAssignment(
                show_id=show_id,
                shift_id=shift.id,
                volunteer_id=vol.id,
                status=a_status,
            )
            db.add(a)
            stats["assignments_created"] += 1

    db.commit()
    return stats


def _parse_bool(val) -> bool:
    if isinstance(val, bool):
        return val
    s = str(val).strip().lower()
    return s in ("true", "1", "yes", "t")


def _parse_xlsx(raw: bytes) -> list:
    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(h).strip().lower() if h else "" for h in next(rows_iter)]
    result = []
    for row in rows_iter:
        d = {}
        for i, val in enumerate(row):
            if i < len(headers) and headers[i]:
                d[headers[i]] = val if val is not None else ""
        result.append(d)
    return result


def _parse_csv(raw: bytes) -> list:
    import csv
    text = raw.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    return list(reader)


# ── Printable HTML Schedule ────────────────────────────────────────

@router.get("/printable")
def printable_schedule(show_id: str, db: Session = Depends(get_db)):
    show = _require_show(db, show_id)
    roles = {r.id: r.name for r in db.execute(select(VolunteerRole).where(VolunteerRole.show_id == show_id)).scalars().all()}
    shifts = db.execute(
        select(VolunteerShift).where(VolunteerShift.show_id == show_id)
        .order_by(VolunteerShift.start_time)
    ).scalars().all()
    assignments = db.execute(select(VolunteerAssignment).where(VolunteerAssignment.show_id == show_id)).scalars().all()
    vols = {v.id: v for v in db.execute(select(Volunteer).where(Volunteer.show_id == show_id)).scalars().all()}

    shift_assigns = {}
    for a in assignments:
        shift_assigns.setdefault(a.shift_id, []).append(a)

    html = [
        "<!DOCTYPE html><html><head><meta charset='utf-8'>",
        "<style>",
        "body{font-family:Arial,sans-serif;margin:20px}",
        "h1{font-size:18px}h2{font-size:14px;margin-top:16px}",
        "table{border-collapse:collapse;width:100%;margin-bottom:12px}",
        "th,td{border:1px solid #333;padding:4px 8px;font-size:12px;text-align:left}",
        "th{background:#eee}",
        "@media print{body{margin:0}h1{font-size:16px}}",
        "</style></head><body>",
        f"<h1>Volunteer Schedule &mdash; {show.name}</h1>",
    ]

    for shift in shifts:
        rname = roles.get(shift.role_id, "Unknown")
        html.append(
            f"<h2>{rname}: {shift.start_time.strftime('%Y-%m-%d %H:%M')}&ndash;{shift.end_time.strftime('%H:%M')}</h2>"
        )
        html.append("<table><tr><th>#</th><th>Name</th><th>Phone</th><th>Status</th></tr>")
        assigned = shift_assigns.get(shift.id, [])
        if assigned:
            for i, a in enumerate(assigned, 1):
                v = vols.get(a.volunteer_id)
                html.append(
                    f"<tr><td>{i}</td><td>{v.name if v else '?'}</td>"
                    f"<td>{v.phone or '' if v else ''}</td>"
                    f"<td>{a.status}</td></tr>"
                )
        else:
            html.append(f"<tr><td colspan='4' style='text-align:center'>No assignments (capacity: {shift.capacity})</td></tr>")
        html.append("</table>")

    html.append("</body></html>")
    return Response(content="".join(html), media_type="text/html")
